import io
import openpyxl
from datetime import datetime
from typing import List, Dict, Any
from src.parsers.base import BaseParser
from src.core.logger import setup_logger

logger = setup_logger(__name__)

class XlsxParser(BaseParser):
    def parse(self, raw_data: bytes, metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Parse Excel workbook (.xlsx) bytes, converting rows into semantic chunks."""
        source_name = metadata.get("source", "Unknown Excel File")
        timestamp = datetime.now().isoformat()
        pages = []
        try:
            file_like = io.BytesIO(raw_data)
            wb = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Fetch all rows as list of cell tuples
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                
                # Identify header row (first non-empty row)
                headers = []
                header_idx = -1
                for r_idx, r in enumerate(rows):
                    cells = [str(c).strip() for c in r if c is not None]
                    if any(cells):
                        headers = [str(c).strip() for c in r if c is not None]
                        header_idx = r_idx
                        break
                        
                if not headers:
                    continue
                    
                # Identify a potential primary key column (default to first column)
                pk_col = None
                for h in headers:
                    if "id" in h.lower() or "key" in h.lower() or "code" in h.lower():
                        pk_col = h
                        break
                if not pk_col and headers:
                    pk_col = headers[0]
                    
                for i, row in enumerate(rows[header_idx + 1:]):
                    row_values = [str(cell).strip() if cell is not None else "" for cell in row]
                    if not any(row_values):
                        continue
                        
                    # Extract primary key value
                    pk_val = None
                    if pk_col and pk_col in headers:
                        try:
                            pk_idx = headers.index(pk_col)
                            pk_val = row_values[pk_idx] if pk_idx < len(row_values) else f"Row_{i+1}"
                        except Exception:
                            pk_val = f"Row_{i+1}"
                    else:
                        pk_val = f"Row_{i+1}"
                        
                    # Build semantic text representation
                    text_lines = [
                        f"Table: {sheet_name} (Workbook: {source_name})",
                        "Column Headers: " + " | ".join(headers),
                        f"Row Data (Primary Key {pk_col or 'Index'} = {pk_val}):"
                    ]
                    for col_name, val in zip(headers, row_values):
                        text_lines.append(f"- {col_name}: {val}")
                        
                    row_text = "\n".join(text_lines)
                    
                    row_metadata = metadata.copy()
                    row_metadata.update({
                        "source_type": "structured_row",
                        "table_name": sheet_name,
                        "primary_key_column": pk_col or "Index",
                        "primary_key_value": pk_val,
                        "columns": headers,
                        "row_number": i + 1,
                        "sheet_name": sheet_name,
                        "timestamp": timestamp
                    })
                    
                    pages.append({
                        "text": row_text,
                        "metadata": row_metadata
                    })
            logger.info(f"Successfully chunked {len(pages)} rows from Excel workbook {source_name}")
        except Exception as e:
            logger.error(f"Error parsing Excel workbook {source_name}: {e}")
        return pages
